import os
import requests
from columns_enum import *
from openpyxl import Workbook
from openpyxl import load_workbook


def download_file(url, save_path):
    response = requests.get(url)
    try:
        with open(save_path, 'wb') as file:
            file.write(response.content)
    except FileNotFoundError:
        print("Plik nie został znaleziony.")
    except IOError:
        print("Wystąpił błąd podczas operacji na pliku.")
    else:
        print(f"Plik pobrany pomyślnie.")


def load_data(xlsx_file):
    workbook = load_workbook(filename=xlsx_file)
    sheets_list = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        sheets_list.append({sheet_name: rows})

    return sheets_list


def sort_column_alphabetically(rows, column_nums):
    sorted_rows = []
    for row in rows:
        sorted_row = list(row) #copy of row to modify
        for column_num in column_nums:
            column_values = row[column_num]

            if len(column_values) > 1:
                column_text = str(column_values)
                words = column_text.split('\n')
                words.sort()
                sorted_string = '\n'.join(words)
                sorted_row[column_num] = sorted_string #modified row
        sorted_rows.append(sorted_row)     
    return sorted_rows


def capitalize_column_values(rows, column_nums):
    capitalized_rows = []

    for i, row in enumerate(rows):
        capitalized_row = list(row) #copy of row to modify
        if i == 0:
            capitalized_rows.append(row)
        else:
            for column_num in column_nums:
                capitalized_row[column_num] = str(row[column_num]).title() #modified row
            capitalized_rows.append(capitalized_row)
    return capitalized_rows

def clear_white_space(rows):
    clean_rows = []
    for row in rows:
        clean_row = list(row) #copy of row to mdidy
        for cell in row:
            cleaned = str(cell).strip()
            clean_row.append(cleaned)
        clean_rows.append(clean_row)
    return clean_rows

def extract_year(file, sheet_name='Intro'):
    year = int(extract_rows(file,sheet_name)[0][-1])
    return year

def extract_rows(file, sheet_name):
    rows= file[sheet_name]
    return rows
    
def show_changes(file_1, file_2, column_num, sheet_name):
    rows_file_1 = extract_rows(file_1,sheet_name) 
    rows_file_2 = extract_rows(file_2,sheet_name) 

    new_in_file_2 = []
    missing_in_file_2 = []
    changed_in_file_2 = []

    values_file_1 = []
    for row in rows_file_1:
        value = row[column_num]
        values_file_1.append(value)

    values_file_2 = []
    for row in rows_file_2:
        value = row[column_num]
        values_file_2.append(value)

    ids_file_2 = []
    for row in rows_file_2:
        id = row[0]
        ids_file_2.append(id)

    #find missing id in file 2
    for value in values_file_1:
        if value not in values_file_2:
            missing_in_file_2.append(value)

    #find new id in file 2
    for value in values_file_2:
        if value not in values_file_1:
            new_in_file_2.append(value)

    #find changed records in file 2 and return their ids
    for i, value in enumerate(values_file_2):
        if value not in values_file_1:
            changed_in_file_2.append(ids_file_2[i])

    changes = {
        'new_records': new_in_file_2,
        'missing_records': missing_in_file_2,
        'changed_records_ids' : changed_in_file_2
    }

    return changes

def get_changes_dict(file_1, file_2, enum, sheet_name): 
    changes_dict = {}
    for column in enum:
        changes_dict[column.name] = show_changes(file_1, file_2, column.value, sheet_name)
    return changes_dict

def generate_report(file_1, file_2, enum, sheet_name):
    list_of_changes=[]
    temp = ''
    if sheet_name == 'Elektronarzedzia':
        temp = 'elektronarzędzi'
    if sheet_name == 'Ostrza':
        temp = 'ostrzy'

    changes_dict = get_changes_dict(file_1, file_2, enum, sheet_name)
    discontinued = changes_dict['ID']['missing_records']
    new_products = changes_dict['ID']['new_records']

    modified = []
    for column_name, changes in changes_dict.items():
        if column_name == 'ID':
            continue
        if changes['changed_records_ids']:
            modified.append(f"Kolumna {column_name} zmieniła się dla {temp}: {changes['changed_records_ids']}")
    
    if discontinued:
        discontinued_str = "Narzędzia wycofane z oferty: " + ", ".join(discontinued) 
        list_of_changes.append(discontinued_str)
    if new_products:
        new_products_str = "Nowe produkty: " + ", ".join(new_products) 
        list_of_changes.append(new_products_str)
    if modified:
        for v in modified:
            list_of_changes.append(v)
    return list_of_changes

def save_data(data, output_file):
    if not os.path.exists("processed"):
        os.makedirs("processed")
    workbook = Workbook()

    for sheet_name, rows in data.items():
        if rows:
            print(f"Saving {len(rows)} rows for sheet: {sheet_name}")
            worksheet = workbook.create_sheet(title=sheet_name)
            for row in rows:
                if isinstance(row, list):  #check if row is a list
                    worksheet.append(row)
                else:
                    print(f"Skipping row in {sheet_name}: {row}. It's not in the expected format.")
        else:
            print(f"No data found for sheet: {sheet_name}")

    workbook.save(filename=output_file)
    

def merge_sheet_info(file_1, file_2, sheet_name, enum, enum_rep):
    rows_1 = extract_rows(file_1, sheet_name)[1:]
    rows_2 = extract_rows(file_2, sheet_name)[1:]
    changes_dict = get_changes_dict(file_1, file_2, enum, sheet_name)
    processed_ids = []
    merged_rows = []

    temp_row = []
    for column in enum_rep:
        temp_row.append(column.name)
    merged_rows.append(temp_row)

    temp_row = []
    for missing_id in changes_dict['ID']['missing_records']:
        temp_row = ['wycofane', missing_id]
        processed_ids.append(missing_id)
        for _ in range(len(enum_rep) - 1):
            temp_row.append(None)  # Append None for the rest of the columns
        merged_rows.append(temp_row)

    temp_row = []
    for new_id in changes_dict['ID']['new_records']:
        temp_row = ['nowe', new_id]
        processed_ids.append(new_id)
        for _ in range(len(enum_rep) - 1):
            temp_row.append(None)  # Append None for the rest of the columns
        merged_rows.append(temp_row)
    
    temp_row = []
    for column_name, changes in changes_dict.items():
        most_changes = 0
        if len(changes['changed_records_ids']) > most_changes:
            most_changes = changes['changed_records_ids']
            most_changes_column = column_name
    
    temp_row = []
    for changed_id in changes_dict[most_changes_column]['changed_records_ids']:
        temp_row = ['zmodyfikowane', changed_id]
        processed_ids.append(changed_id)
        for _ in range(len(enum_rep) - 1):
            temp_row.append(None)  # Append None for the rest of the columns
        merged_rows.append(temp_row)
    
    temp_row = []
    for row in rows_2:
        id = row[0]
        if id not in processed_ids:
            temp_row = ['-', id]
        else :
            continue
        merged_rows.append(temp_row)    
    
    return merged_rows

    # temp_row = [None] * len(enum_rep)
    # for column_r in enum_rep:
    #     for column in enum:
    #         if column.name == "ID" :
    #             temp_row[column_r.ID.value] = rows_2[0][0]
    #         if "23" in column_r.name and column.name in column_r.name:
    #             for cell in rows_1:
    #                 temp_row[column_r.value] = cell[column.value]
                        
    #         elif "24" in column_r.name and column.name in column_r.name:
    #             for cell in rows_2:
    #                 temp_row[column_r.value] = cell[column.value]
    #     merged_rows.append(temp_row)

            